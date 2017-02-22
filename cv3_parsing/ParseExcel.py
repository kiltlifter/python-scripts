#!/usr/bin/python
# -*- coding: utf-8 -*-
from __future__ import print_function
from openpyxl import load_workbook
import warnings
import re

__author__ = "sean douglas"


class Parser:
	def __init__(self, filename, features_sheet=None, groups_sheet=None):
		warnings.simplefilter("ignore")
		self.wb = load_workbook(filename)
		self.features_sheet = features_sheet
		self.groups_sheet = groups_sheet

	def cd3_groups(self):
		ws = self.wb[self.features_sheet]
		# Adjust for CD of choice
		cd3_range = 'Q2:S61'

		group_rows = set()
		for row in ws[cd3_range]:
			for cell in row:
				if cell.value:
					group_rows.add(cell.row)

		group_set = set()
		for row_val in group_rows:
			group_range = "D{0}:L{1}".format(row_val, row_val)
			for cell in ws[group_range]:
				for c in cell:
					if c.value:
						group_set.add(c.value)

		strip_hyper = lambda x: re.findall("=HYPERLINK\(\"#groups!.*\", \"(Group\d+)\"\)", x)[-1]
		clean_groups = [strip_hyper(x) for x in group_set]
		clean_groups.append("Standards")
		return clean_groups

	def cd2_and_3_groups(self):
		ws = self.wb[self.features_sheet]
		# Adjust for CD of choice
		cd3_range = 'O2:S61'

		group_rows = set()
		for row in ws[cd3_range]:
			for cell in row:
				if cell.value:
					group_rows.add(cell.row)

		group_set = set()
		for row_val in group_rows:
			group_range = "D{0}:L{1}".format(row_val, row_val)
			for cell in ws[group_range]:
				for c in cell:
					if c.value:
						group_set.add(c.value)

		strip_hyper = lambda x: re.findall("=HYPERLINK\(\"#groups!.*\", \"(Group\d+)\"\)", x)[-1]
		clean_groups = [strip_hyper(x) for x in group_set]
		clean_groups.append("Standards")
		return clean_groups

	def requirements_in_cd3(self):
		ws = self.wb[self.groups_sheet]
		groups = self.cd3_groups()
		group_uid_range = "A2:B563"

		uids_in_groups = {}
		for row in ws[group_uid_range]:
			if row[0].value:
				if row[0].value in groups:
					group = {"count": None, "requirements": []}
					uid_range = "B{0}:B{1}".format(row[0].row+1, row[0].row+100)
					for uid in ws[uid_range]:
						if uid[0].value:
							if re.search("^UID", uid[0].value):
								group["requirements"].append(uid[0].value)
							else:
								break
						else:
							break
					group["count"] = len(group["requirements"])
					uids_in_groups[row[0].value] = group
		return uids_in_groups

	def requirements_in_cd2_and_3(self):
		ws = self.wb[self.groups_sheet]
		groups = self.cd2_and_3_groups()
		group_uid_range = "A2:B563"

		uids_in_groups = {}
		for row in ws[group_uid_range]:
			if row[0].value:
				if row[0].value in groups:
					group = {"count": None, "requirements": []}
					uid_range = "B{0}:B{1}".format(row[0].row+1, row[0].row+100)
					for uid in ws[uid_range]:
						if uid[0].value:
							if re.search("^UID", uid[0].value):
								group["requirements"].append(uid[0].value)
							else:
								break
						else:
							break
					group["count"] = len(group["requirements"])
					uids_in_groups[row[0].value] = group
		return uids_in_groups

	def uids_in_cd3(self):
		req_in_cd3 = self.requirements_in_cd3()
		uid_list = []
		for i in req_in_cd3:
			for r in req_in_cd3[i]['requirements']:
				uid_list.append(r)
		return uid_list

	def uids_in_cd2_and_cd3(self):
		req_in_cd2_and_cd3 = self.requirements_in_cd2_and_3()
		uid_list = []
		for i in req_in_cd2_and_cd3:
			for r in req_in_cd2_and_cd3[i]['requirements']:
				uid_list.append(r)
		return uid_list

	def stats(self):
		groups = self.cd3_groups()
		groups2 = self.cd2_and_3_groups()
		requirements = self.requirements_in_cd3()
		requirements2 = self.requirements_in_cd2_and_3()

		print("Count of Groups and Standards in CD3: {}".format(len(groups)))
		print("Count of Groups and Standards in CD2 and CD3: {}".format(len(groups2)))
		print("Sum of requirements in CD3: {}".format(sum([requirements[x]['count'] for x in requirements])))
		print("Sum of requirements in CD2 and CD3: {}".format(sum([requirements2[x]['count'] for x in requirements2])))

		stat_data = [["Count of Groups in CD3:", "{}".format(len(groups))],
		                ["Count of Groups in CD2 and CD3:", "{}".format(len(groups2))],
		                ["Sum of requirements in CD3:", "{}".format(sum([requirements[x]['count'] for x in requirements]))],
						["Sum of requirements in CD2 and CD3:", "{}".format(sum([requirements2[x]['count'] for x in requirements2]))]
			]
		return stat_data

	def prioritized_requirements(self):
		ws = self.wb[self.groups_sheet]
		req_range = "B3:E563"

		cd3_requirements = self.requirements_in_cd3()
		requirements = set()
		for group in [cd3_requirements[r]["requirements"] for r in cd3_requirements]:
			for r in group:
				requirements.add(r)

		requirements_with_priority = {}
		for row in ws[req_range]:
			if row[0].value:
				if row[0].value in requirements:
					pri = str(row[3].value) if row[3].value else None
					requirements_with_priority[row[0].value] = [row[1].value, row[2].value, pri]
		return requirements_with_priority

	def prioritized_requirements_cd2_and_cd3(self):
		ws = self.wb[self.groups_sheet]
		req_range = "B3:E563"

		cd3_requirements = self.requirements_in_cd2_and_3()
		requirements = set()
		for group in [cd3_requirements[r]["requirements"] for r in cd3_requirements]:
			for r in group:
				requirements.add(r)

		requirements_with_priority = {}
		for row in ws[req_range]:
			if row[0].value:
				if row[0].value in requirements:
					pri = str(row[3].value) if row[3].value else None
					requirements_with_priority[row[0].value] = [row[1].value, row[2].value, pri]
		return requirements_with_priority

	def groups_with_added_data(self):
		ws = self.wb[self.features_sheet]
		groups = self.cd3_groups()
		data_range = "A2:L61"
		strip_hyper = lambda x: re.findall("=HYPERLINK\(\"#groups!.*\", \"(Group\d+)\"\)", x)[-1]
		strip_hyper2 = lambda x: re.findall("=HYPERLINK\(\"#activities!.*\", \"(.*)\"\)", x)[-1]

		enriched_group_data = {}
		for row in ws[data_range]:
			for cell in row[3:]:
				if cell.value and strip_hyper(cell.value) in groups:
					if strip_hyper(cell.value) in enriched_group_data.keys():
						enriched_group_data[strip_hyper(cell.value)][0] += "\n{}".format(strip_hyper2(row[1].value))
						if not re.search(row[2].value, enriched_group_data[strip_hyper(cell.value)][1]):
							enriched_group_data[strip_hyper(cell.value)][1] += "\n{}".format(row[2].value)
					else:
						enriched_group_data[strip_hyper(cell.value)] = [strip_hyper2(row[1].value), row[2].value]
		return enriched_group_data

	def cd2_and_cd3_groups_with_added_data(self):
		ws = self.wb[self.features_sheet]
		groups = self.cd2_and_3_groups()
		data_range = "A2:L61"
		strip_hyper = lambda x: re.findall("=HYPERLINK\(\"#groups!.*\", \"(Group\d+)\"\)", x)[-1]
		strip_hyper2 = lambda x: re.findall("=HYPERLINK\(\"#activities!.*\", \"(.*)\"\)", x)[-1]

		enriched_group_data = {}
		for row in ws[data_range]:
			for cell in row[3:]:
				if cell.value and strip_hyper(cell.value) in groups:
					if strip_hyper(cell.value) in enriched_group_data.keys():
						enriched_group_data[strip_hyper(cell.value)][0] += "\n{}".format(strip_hyper2(row[1].value))
						if not re.search(row[2].value, enriched_group_data[strip_hyper(cell.value)][1]):
							enriched_group_data[strip_hyper(cell.value)][1] += "\n{}".format(row[2].value)
					else:
						enriched_group_data[strip_hyper(cell.value)] = [strip_hyper2(row[1].value), row[2].value]
		return enriched_group_data


