#!/usr/bin/python
# -*- coding: utf-8 -*-
from __future__ import print_function
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Fill
import ParseExcel

__author__ = "sean douglas"


class Write:
	def __init__(self, filename, requirements_sheet=None):
		self.filename = filename
		self.wb = load_workbook(filename)
		self.requirements_sheet = requirements_sheet

	def write_sheet(self, data):
		requirements_sheet = self.wb.create_sheet("CD3 Requirements")
		header = ("REQUIREMENT", "PARENT", "DESCRIPTION", "PRIORITY", "GROUP", "GROUP IN FEATURE(s)", "OWNER")

		requirements_sheet.append(header)
		for row in data:
			requirements_sheet.append(row)

	def write_cd2_and_cd3_sheet(self, data):
		requirements_sheet = self.wb.create_sheet(self.requirements_sheet)
		header = ("REQUIREMENT", "PARENT", "DESCRIPTION", "PRIORITY", "GROUP", "GROUP IN FEATURE(s)", "OWNER")

		requirements_sheet.append(header)
		for row in data:
			requirements_sheet.append(row)

	def write_stats_to_sheet(self, data):
		ws = self.wb[self.requirements_sheet]
		# Add two blank rows
		ws.append(("", "", "", "", "", "", ""))
		ws.append(("", "", "", "", "", "", ""))
		ws.append(data)

	def highlight_cd2_not_in_cd3_requirements(self, requirement_uids):
		fill = PatternFill("solid", fgColor="ADD8E6")
		ws = self.wb[self.requirements_sheet]
		for row in ws["A2:A600"]:
			if row[0].value:
				if row[0].value not in requirement_uids:
					print(row[0].value)
					row[0].fill = fill

	def highlight_cd2_in_cd3_requirements(self, requirement_uids):
		fill = PatternFill("solid", fgColor="FFFFFF")
		ws = self.wb[self.requirements_sheet]
		for row in ws["A2:A600"]:
			if row[0].value:
				if row[0].value in requirement_uids:
					print(row[0].value)
					row[0].fill = fill

	def write_sheet_data(self):
		self.wb.save(self.filename)
